import os.path
import time


import csv
from typing import Any, Dict, List

from openpyxl import Workbook

from app.helpers import sanitize_pet_field, get_logger, get_timedate_string

logger = get_logger(__name__)

current_dir = os.path.dirname(__file__)
parent_dir = os.path.split(current_dir)[0]
files_dir = os.path.join(current_dir, "files")

OUTPUT_FILE_NAME = "data"


class WriteController(object):
    resource_name: str
    headers: List[str]

    def write_csv(self, data: list):
        if not self._csv_file_exists():
            self.create_csv_file()

        try:
            with open(self._get_csv_file(), "a", newline="") as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=self.headers)
                logger.debug(f"Writing {len(data)} rows")

                for d in data:
                    try:
                        writer.writerow(d)
                    except UnicodeEncodeError as e:
                        _d = {}
                        for key, value in d.items():
                            if key == "name":
                                key = f"_{key}"
                            if isinstance(value, str):
                                _d[key] = value.encode("utf-8").decode(
                                    "ascii", "ignore"
                                )
                            else:
                                _d[key] = value

                        logger.error(f"{e} - {_d}")

                logger.info(f"Successfully wrote {len(data)} rows")
        except PermissionError:
            time.sleep(1)
            self.write_csv(data)

    def _get_csv_file(self):
        return os.path.join(
            files_dir,
            f"{self.resource_name}_{get_timedate_string()}.csv",
        )

    def _csv_file_exists(self):
        return os.path.exists(self._get_csv_file())

    def create_csv_file(self):
        with open(self._get_csv_file(), "a", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(self.headers)

    def write(self, data: list):
        wb = self._get_workbook()

        try:
            sheet = wb[self.resource_name]
        except KeyError:
            sheet = wb.create_sheet(self.resource_name)
            sheet.append(list(data[0].keys()))  # set headers

        logger.info(f"Writing {len(data)} rows")

        for row in data:
            sheet.append(list(row.values()))  # write row

        wb.save(
            os.path.join(
                files_dir, f"{OUTPUT_FILE_NAME}_{get_timedate_string()}.xlsx"
            )
        )

        logger.info(f"Successfully wrote {len(data)} rows")

    def _workbook_exists(self):
        if os.path.exists(
            os.path.join(
                files_dir, f"{OUTPUT_FILE_NAME}_{get_timedate_string()}.xlsx"
            )
        ):
            return True
        else:
            return False

    def _remove_untitled_sheet(self, workbook):
        untitled_ws = workbook["Sheet"]
        workbook.remove(untitled_ws)

    def _get_workbook(self):
        if self._workbook_exists():
            wb = Workbook(
                os.path.join(
                    files_dir,
                    f"{OUTPUT_FILE_NAME}_{get_timedate_string()}.xlsx",
                ),
            )
        else:
            wb = Workbook(write_only=True)
            # self._remove_untitled_sheet(wb)

        return wb


class AddressesWriteController(WriteController):
    """Controller class used for writing addresses."""

    resource_name = "addresses"
    headers = [
        "id",
        "pimsId",
        "address1",
        "address2",
        "city",
        "state",
        "postalCode",
        "clientId",
        "patientIds",
    ]

    def extract_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract data needed for writing from a dictionary (rj)."""
        ret = {
            "id": data["id"],
            "pimsId": data["pimsId"],
            "address1": data["addressLine1"],
            "address2": data["addressLine2"],
            "city": data["city"],
            "state": data["stateProvince"],
            "postalCode": data["postalCode"],
        }

        if data.get("client"):
            ret.update(
                {
                    "clientId": data["client"]["pimsId"],
                }
            )
            if data["client"].get("pets"):
                ret.update(
                    {
                        "patientIds": sanitize_pet_field(
                            data["client"]["pets"], "pimsId"
                        ),
                    }
                )

        return ret


class BreedsWriteController(WriteController):
    resource_name = "breeds"
    headers = [
        "id",
        "pimsId",
        "description",
    ]

    def extract_data(self, data):
        return {
            "id": data["id"],
            "pimsId": data["pimsId"],
            "description": data["description"],
        }


class SpeciesWriteController(WriteController):
    resource_name = "species"
    headers = [
        "id",
        "pimsId",
        "description",
    ]

    def extract_data(self, data):
        return {
            "id": data["id"],
            "pimsId": data["pimsId"],
            "description": data["description"],
        }


class ClientsWriteController(WriteController):
    resource_name = "clients"
    headers = [
        "id",
        "firstName",
        "lastName",
        "balance",
        "clientClassificationCode",
        "clientClassificationName",
        "patientIds",
        "createdAt",
    ]

    def extract_data(self, data):
        return {
            "id": data["pimsId"],
            "firstName": data["firstName"],
            "lastName": data["lastName"],
            "balance": data["balance"],
            "clientClassificationCode": data["classificationCode"],
            "clientClassificationName": data["classificationDescription"],
            "patientIds": sanitize_pet_field(data["pets"], "pimsId"),
            "createdAt": data["enteredDate"],  # enteredAt
        }


class CodesWriteController(WriteController):
    resource_name = "codes"
    headers = [
        "pimsId",
        "id",
        "code",
        "codeDescription",
        "codeCategory",
        "codeCategoryDescription",
    ]

    def extract_data(self, data):
        return {
            "pimsId": data["pimsId"],
            "id": data["id"],
            "code": data["code"],
            "codeDescription": data["itemDescription"],
            "codeCategory": data["codeCategory"],
            "codeCategoryDescription": data["codeCategoryDescription"],
        }


class PatientsWriteController(WriteController):
    resource_name = "patients"
    headers = [
        "id",
        "name",
        "color",
        "currentWeight",
        "dateOfBirth",
        "dateOfDeath",
        "genderDescription",
        "microchip",
        "rabiesId",
        "allergies",
        "alerts",
        "createdAt",
        "clientId",
        "breedId",
        "breedPimsId",
        "speciesId",
        "speciesPimsId",
    ]

    def extract_data(self, data):
        ret = {
            "id": data["pimsId"],
            "name": data["name"],
            # "breedId": data[''], relationship with `breed.id`
            # "speciesId": data[''], relationship with `species.id`
            "color": data["color"],
            "currentWeight": data["currentWeight"],
            "dateOfBirth": data["dateOfBirth"],
            "dateOfDeath": data["dateOfDeath"],
            "genderDescription": data["genderDescription"],
            "microchip": data["microChip"],
            "rabiesId": data["rabies"],
            "allergies": data["allergies"],
            "alerts": data["alerts"],
            # "patientWeightHistory": data[''],
            # "weight": data[''], how is this different than currentWeight
            "createdAt": data["enteredDate"],
        }

        if data.get("owner"):
            ret.update(
                {
                    "clientId": data["owner"]["pimsId"],
                }
            )

        if data.get("patientBreed"):
            ret.update(
                {
                    "breedId": data["patientBreed"]["id"],
                    "breedPimsId": data["patientBreed"]["pimsId"],
                }
            )

        if data.get("patientSpecies"):
            ret.update(
                {
                    "speciesId": data["patientSpecies"]["id"],
                    "speciesPimsId": data["patientSpecies"]["pimsId"],
                }
            )

        return ret


class PhoneNumbersWriteController(WriteController):
    resource_name = "phoneNumbers"
    headers = [
        "pimsId",
        "number",
        "isSMSEnabled",
        "isPrimary",
        "clientId",
    ]

    def extract_data(self, data):
        return {
            "pimsId": data["pimsId"],
            "number": data["number"],
            "isSMSEnabled": data["sMS"],
            "isPrimary": data["isPrimary"],
            "clientId": data["client"]["pimsId"],
        }


class SOAPsWriteController(WriteController):
    resource_name = "sOAPs"
    headers = [
        "pimsId",
        "subjective",
        "objective",
        "assessment",
        "plan",
        "patientId",
        "createdAt",
    ]

    def extract_data(self, data):
        return {
            "pimsId": data["pimsId"],
            "subjective": data["subjective"],
            "objective": data["objective"],
            "assessment": data["assessment"],
            "plan": data["plan"],
            "patientId": data["patient"]["pimsId"],
            "createdAt": data["createDate"],
        }


class TransactionsWriteController(WriteController):
    resource_name = "transactions"
    headers = [
        "pimsId",
        "amount",
        "invoiceNumber",
        "quantity",
        "transactionDate",
        "comments",
        "description",
        "clientId",
        "patientId",
        "siteId",
        "sitePimsId",
    ]

    def extract_data(self, data):
        ret = {
            "pimsId": data["pimsId"],
            "amount": data["amount"],
            "invoiceNumber": data["invoiceNumber"],
            "quantity": data["quantity"],
            "transactionDate": data["datePerformed"],
            "comments": data["comments"],
            "description": data["description"],
        }

        if data.get("client"):
            ret.update(
                {
                    "clientId": data["client"]["pimsId"],
                }
            )

        if data.get("patient"):
            ret.update(
                {
                    "patientId": data["patient"]["pimsId"],
                }
            )

        if data.get("site"):
            ret.update(
                {
                    "siteId": data["site"]["id"],
                    "sitePimsId": data["site"]["pimsId"],
                }
            )

        return ret


class VaccinesWriteController(WriteController):
    resource_name = "vaccines"
    headers = [
        "pimsId",
        "vaccineName",
        "inventoryItemDescription",
        "manufactrer",
        "tagNumber",
        "dateVaccinated",
        "vaccineExpiration",
        "patientId",
        "clientFirstName",
        "clientLastName",
        "clientId",
    ]

    def extract_data(self, data):
        return {
            "pimsId": data["pimsId"],
            "vaccineName": data["vaccineDescription"],
            "inventoryItemDescription": data["description"],
            "manufactrer": data["manufacturer"],
            "tagNumber": data["tag"],
            "dateVaccinated": data["dateGiven"],
            "vaccineExpiration": data["expirationDate"],
            "patientId": data["patient"]["pimsId"],
            "clientFirstName": data["patient"]["owner"]["firstName"],
            "clientLastName": data["patient"]["owner"]["lastName"],
            "clientId": data["patient"]["owner"]["pimsId"],
        }
